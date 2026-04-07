"""Excel reading helpers for ingestion steps.

Usage:
    from ingestion.excel_reader import read_entities_workbook, read_sources_workbook
    entities_rows, entity_type_rows = read_entities_workbook("path/to/workbook.xlsx")
    sources_rows, source_assets_rows = read_sources_workbook("path/to/workbook.xlsx")

This module does not perform validation or database writes.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _is_empty_cell(value: Any) -> bool:
    """Return True when a worksheet cell value should be treated as empty.

    Empty values include:
    - None
    - strings that are blank/whitespace-only
    """
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def worksheet_to_dict_rows(worksheet: Worksheet) -> list[dict[str, Any]]:
    """Convert a worksheet into list[dict] using the first row as headers.

    Fully empty data rows are skipped.
    Header names are preserved exactly as they appear in row 1.

    Args:
        worksheet: OpenPyXL worksheet.

    Returns:
        List of dictionaries where keys are header values and values are cell values.
    """
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = list(rows[0])
    parsed_rows: list[dict[str, Any]] = []

    for row in rows[1:]:
        if all(_is_empty_cell(cell) for cell in row):
            continue

        parsed: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if header is None:
                continue
            parsed[str(header)] = row[idx] if idx < len(row) else None
        parsed_rows.append(parsed)

    return parsed_rows


def read_entities_workbook(workbook_path: str | Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Read entities_seed + entity_types sheets from workbook.

    This function expects:
    - ``entities_seed`` sheet
    - ``entity_types`` sheet

    Returns:
        Tuple of (entities_rows, entity_type_rows), where each element is
        ``list[dict[str, Any]]``.

    Raises:
        ValueError: when required sheets are missing.
    """
    workbook = load_workbook(filename=Path(workbook_path), data_only=True, read_only=True)
    required = ("entities_seed", "entity_types")
    missing = [sheet for sheet in required if sheet not in workbook.sheetnames]
    if missing:
        raise ValueError(
            f"Workbook is missing required sheet(s): {', '.join(missing)}. "
            f"Required: {', '.join(required)}."
        )

    entities_rows = worksheet_to_dict_rows(workbook["entities_seed"])
    entity_type_rows = worksheet_to_dict_rows(workbook["entity_types"])
    return entities_rows, entity_type_rows


def read_sources_workbook(workbook_path: str | Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Read sources_registry + source_assets sheets from workbook.

    This function expects:
    - ``sources_registry`` sheet
    - ``source_assets`` sheet

    Returns:
        Tuple of (sources_rows, source_assets_rows), where each element is
        ``list[dict[str, Any]]``.

    Raises:
        ValueError: when required sheets are missing.
    """
    workbook = load_workbook(filename=Path(workbook_path), data_only=True, read_only=True)
    required = ("sources_registry", "source_assets")
    missing = [sheet for sheet in required if sheet not in workbook.sheetnames]
    if missing:
        raise ValueError(
            f"Workbook is missing required sheet(s): {', '.join(missing)}. "
            f"Required: {', '.join(required)}."
        )

    sources_rows = worksheet_to_dict_rows(workbook["sources_registry"])
    source_assets_rows = worksheet_to_dict_rows(workbook["source_assets"])
    return sources_rows, source_assets_rows
