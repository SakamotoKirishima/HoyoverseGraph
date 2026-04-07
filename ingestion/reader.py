"""Workbook reader utilities for ontology ingestion.

This module reads selected ontology sheets from an Excel workbook and converts
rows into ``list[dict]`` records keyed by original header names.

Run (module usage via CLI entrypoint):
    python -m ingestion.main <path-to-workbook.xlsx-or-xlsm>
Direct usage (Python):
    from ingestion.reader import read_ontology_workbook
    data = read_ontology_workbook("docs/hoyoverse_ontology_v1.xlsm")
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

SHEETS_TO_PARSE: tuple[str, ...] = (
    "entity_types",
    "relationship_types",
    "evidence_labels",
    "entities_seed",
    "claims_seed",
    "sources_registry",
    "source_assets",
    "editorial_rules",
)

REQUIRED_SHEETS: tuple[str, ...] = (
    "entities_seed",
    "claims_seed",
    "sources_registry",
    "source_assets",
)

SOURCES_REGISTRY_HEADERS: set[str] = {"source_id", "title", "source_type", "source_format"}


def _is_empty_cell(value: Any) -> bool:
    """Return True when a worksheet cell value should be treated as empty."""
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _find_header_row_index(
    worksheet: Worksheet,
    expected_headers: set[str],
    max_scan_rows: int = 25,
) -> int:
    """Find 1-based header row index by matching expected header names.

    Falls back to row 1 when no matching row is found.
    """
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        header_candidates = {
            cell.strip()
            for cell in row
            if isinstance(cell, str) and cell.strip() != ""
        }
        if expected_headers.issubset(header_candidates):
            return row_index
        if row_index >= max_scan_rows:
            break
    return 1


def parse_worksheet_rows(
    worksheet: Worksheet,
    *,
    header_row_index: int = 1,
    stop_at_first_empty_row: bool = False,
) -> list[dict[str, Any]]:
    """Parse one worksheet into row dictionaries using a header row.

    Header names are preserved exactly as found in the header row.
    Rows that are fully empty are skipped. When stop_at_first_empty_row is True,
    parsing stops at the first fully empty row after data begins.
    """
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    if header_row_index < 1 or header_row_index > len(rows):
        return []

    header_row = rows[header_row_index - 1]
    headers = list(header_row)
    parsed_rows: list[dict[str, Any]] = []
    started_data = False

    for row in rows[header_row_index:]:
        is_empty_row = all(_is_empty_cell(cell) for cell in row)
        if is_empty_row:
            if stop_at_first_empty_row and started_data:
                break
            continue

        row_dict: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if header is None:
                continue
            cell_value = row[index] if index < len(row) else None
            row_dict[str(header)] = cell_value

        if row_dict:
            parsed_rows.append(row_dict)
            started_data = True

    return parsed_rows


def read_ontology_workbook(workbook_path: str | Path) -> dict[str, list[dict[str, Any]]]:
    """Read selected ontology sheets from a workbook file.

    Args:
        workbook_path: Path to an .xlsx/.xlsm workbook.

    Returns:
        A mapping from sheet name to parsed rows (list of dicts).
        Only known sheets in SHEETS_TO_PARSE are included, and only when present.

    Raises:
        ValueError: If any required sheets are missing.
    """
    workbook = load_workbook(filename=Path(workbook_path), data_only=True, read_only=True)
    available_sheet_names = set(workbook.sheetnames)

    missing_required = [sheet for sheet in REQUIRED_SHEETS if sheet not in available_sheet_names]
    if missing_required:
        missing_display = ", ".join(missing_required)
        raise ValueError(
            "Workbook is missing required sheet(s): "
            f"{missing_display}. Required: {', '.join(REQUIRED_SHEETS)}."
        )

    parsed: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in SHEETS_TO_PARSE:
        if sheet_name not in available_sheet_names:
            continue
        worksheet = workbook[sheet_name]
        header_row_index = 1
        if sheet_name == "sources_registry":
            header_row_index = _find_header_row_index(
                worksheet, SOURCES_REGISTRY_HEADERS
            )
        parsed[sheet_name] = parse_worksheet_rows(
            worksheet,
            header_row_index=header_row_index,
            stop_at_first_empty_row=True,
        )

    return parsed
